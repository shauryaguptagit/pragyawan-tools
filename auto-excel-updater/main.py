import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
import openpyxl
import os
from datetime import datetime
import time
import re

class ExcelUpdaterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Auto Excel Updater V2.1")
        self.root.geometry("800x700")
        self.activity_log = []
        
        # Configure layout
        main_frame = ttk.Frame(root, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Input section
        input_frame = ttk.LabelFrame(main_frame, text="Input Data", padding=10)
        input_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(input_frame, text="Input (Keyword:Value format):").grid(row=0, column=0, sticky="w")
        self.input_text = tk.Text(input_frame, height=8, width=80)
        self.input_text.grid(row=1, column=0, columnspan=2, pady=(0, 10), sticky="ew")
        
        # File path
        ttk.Label(input_frame, text="Excel File Path:").grid(row=2, column=0, sticky="w")
        self.file_path = ttk.Entry(input_frame, width=60)
        self.file_path.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        
        browse_btn = ttk.Button(input_frame, text="Browse", command=self.browse_file)
        browse_btn.grid(row=3, column=1, padx=(10, 0))
        
        # Options
        self.backup_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(input_frame, text="Create backup before updating", variable=self.backup_var
                       ).grid(row=4, column=0, sticky="w", columnspan=2)
        
        # Process button
        self.process_btn = ttk.Button(input_frame, text="Update Excel", command=self.process_input)
        self.process_btn.grid(row=5, column=0, columnspan=2, pady=10)
        
        # Activity Log
        log_frame = ttk.LabelFrame(main_frame, text="Activity Log", padding=10)
        log_frame.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = scrolledtext.ScrolledText(
            log_frame, height=15, state=tk.DISABLED, wrap=tk.WORD
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        # Clear log button
        clear_btn = ttk.Button(log_frame, text="Clear Log", command=self.clear_log)
        clear_btn.pack(side=tk.RIGHT, pady=(10, 0))
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Set default file path
        self.file_path.insert(0, "Master.xlsx")
        
        # Sample input for testing
        sample_input = """Trade: SC
Location: Jaipur
Dispatch: 5
Inspection: 3"""
        self.input_text.insert("1.0", sample_input)
        
        # Initialize structure cache
        self.structure_cache = {}
        self.last_structure_load = 0
        self.log_activity("Application started")

    def browse_file(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filepath:
            self.file_path.delete(0, tk.END)
            self.file_path.insert(0, filepath)
            # Clear cache when file changes
            self.structure_cache = {}
            self.log_activity(f"Selected file: {os.path.basename(filepath)}")

    def clear_log(self):
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.log_activity("Log cleared")

    def log_activity(self, message):
        """Add timestamped message to activity log"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}"
        self.activity_log.append(log_entry)
        
        # Update log display
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, log_entry + "\n")
        self.log_text.see(tk.END)  # Scroll to bottom
        self.log_text.config(state=tk.DISABLED)
        
        # Update status bar
        self.status_var.set(message)

    def parse_input(self, text):
        """Extract key-value pairs from input text with flexible parsing"""
        data = {}
        for line in text.strip().split('\n'):
            if ':' in line:
                parts = line.split(':', 1)
                key = parts[0].strip()
                value = parts[1].strip()
                data[key] = value
        return data

    def create_backup(self, file_path):
        """Create timestamped backup of the Excel file"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.splitext(file_path)[0] + f"_backup_{timestamp}.xlsx"
        
        import shutil
        shutil.copyfile(file_path, backup_path)
        
        self.log_activity(f"Created backup: {os.path.basename(backup_path)}")
        return backup_path

    def detect_excel_structure(self, sheet):
        """Automatically detect Excel structure including headers and merged cells"""
        structure = {
            'trade_col': None,
            'trade_start_row': None,
            'headers': {},
            'merged_areas': {}
        }
        
        # Detect trade name column (looking for "Trade Name" or similar)
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=3, column=col).value  # Row 3
            if cell_value and "trade" in str(cell_value).lower():
                structure['trade_col'] = col
                structure['trade_start_row'] = 5  # Start at row 5
                break
        
        # If trade column not found, default to column C (3)
        if not structure['trade_col']:
            structure['trade_col'] = 3
            structure['trade_start_row'] = 5
            self.log_activity("Trade column not found, using default (Column C)")

        # Analyze merged cells for headers
        for merged_range in sheet.merged_cells.ranges:
            # Only consider horizontal merges in header rows (rows 3-4)
            if merged_range.min_row in [3, 4] and merged_range.min_row == merged_range.max_row:
                main_header = sheet.cell(merged_range.min_row, merged_range.min_col).value
                if main_header:
                    # Normalize header name for matching
                    normalized_header = self.normalize_header(main_header)
                    
                    # Get sub-headers under the merged area
                    sub_headers = []
                    for col in range(merged_range.min_col, merged_range.max_col + 1):
                        sub_header_cell = sheet.cell(merged_range.min_row + 1, col)
                        if sub_header_cell.value:
                            sub_headers.append({
                                'name': sub_header_cell.value,
                                'col': col
                            })
                    
                    # Store header structure
                    structure['headers'][normalized_header] = {
                        'original_name': main_header,
                        'start_col': merged_range.min_col,
                        'end_col': merged_range.max_col,
                        'sub_headers': sub_headers
                    }
                    structure['merged_areas'][normalized_header] = merged_range
        
        # Log detected headers for debugging
        detected_headers = [info['original_name'] for info in structure['headers'].values()]
        self.log_activity(f"Detected headers: {', '.join(detected_headers)}")
        return structure

    def normalize_header(self, header):
        """Normalize header names for flexible matching"""
        header = str(header).lower().strip()
        # Remove common prefixes/suffixes
        header = re.sub(r'\b(kits?|quantity|total|count)\b', '', header)
        # Remove special characters and extra spaces
        header = re.sub(r'[^\w\s]', '', header)
        header = re.sub(r'\s+', ' ', header).strip()
        
        # Apply special mappings
        if 'inspect' in header:
            return 'inspection'
        if 'dispatch' in header:
            return 'dispatch'
        return header

    def get_excel_structure(self, file_path):
        """Get cached or fresh Excel structure"""
        current_time = time.time()
        if (file_path in self.structure_cache and 
            current_time - self.last_structure_load < 300):  # 5 minute cache
            return self.structure_cache[file_path]
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            sheet = wb["Master Sheet"]
            structure = self.detect_excel_structure(sheet)
            wb.close()
            
            self.structure_cache[file_path] = structure
            self.last_structure_load = current_time
            return structure
        except Exception as e:
            self.log_activity(f"Structure detection error: {str(e)}")
            raise

    def find_target_cell(self, sheet, structure, trade, location, data_type):
        """Find target cell using detected structure"""
        # Find trade row
        trade_row = None
        for row in range(structure['trade_start_row'], sheet.max_row + 1):
            cell = sheet.cell(row, structure['trade_col'])
            if cell.value and trade in str(cell.value):
                trade_row = row
                break
        
        if not trade_row:
            raise ValueError(f"Trade '{trade}' not found in column {structure['trade_col']}")
        
        # Normalize data type for matching
        normalized_type = self.normalize_header(data_type)
        
        # Find matching header for data type
        matching_header = None
        for header_name in structure['headers']:
            if normalized_type in header_name:
                matching_header = header_name
                break
        
        if not matching_header:
            available_headers = [k for k in structure['headers'].keys()]
            raise ValueError(f"No header found matching '{data_type}'. Available: {', '.join(available_headers)}")
        
        # Find location within header's sub-headers
        location_col = None
        for sub_header in structure['headers'][matching_header]['sub_headers']:
            if location.lower() in sub_header['name'].lower():
                location_col = sub_header['col']
                break
        
        if not location_col:
            sub_header_names = [sh['name'] for sh in structure['headers'][matching_header]['sub_headers']]
            raise ValueError(f"Location '{location}' not found under '{matching_header}'. Available: {', '.join(sub_header_names)}")
        
        return trade_row, location_col

    def update_excel(self, file_path, data, create_backup=True):
        """Update Excel file with parsed data"""
        # Create backup if requested
        if create_backup:
            self.create_backup(file_path)
        
        # Load Excel structure
        structure = self.get_excel_structure(file_path)
        
        # Open workbook for updating
        wb = openpyxl.load_workbook(file_path)
        sheet = wb["Master Sheet"]
        
        # Get required parameters from input
        trade_map = {"SC": "Sculptor"}  # Trade abbreviation mapping
        trade_abbr = data.get("Trade", "")
        trade = trade_map.get(trade_abbr, trade_abbr)
        location = data.get("Location", "")
        
        # Process all data types except Trade and Location
        updated_cells = []
        for key, value in data.items():
            if key not in ["Trade", "Location"]:
                data_type = key
                try:
                    # Convert value to numeric
                    value_num = float(value.replace(',', ''))
                    
                    # Find target cell
                    row_idx, col_idx = self.find_target_cell(sheet, structure, trade, location, data_type)
                    
                    # Get current cell value
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    current_value = cell.value
                    
                    # Handle value appending (FIXED: Now appends instead of overwriting)
                    if current_value is None:
                        new_value = value_num
                    else:
                        try:
                            # Try to convert existing value to number
                            current_num = float(current_value)
                            new_value = current_num + value_num
                        except (TypeError, ValueError):
                            # If conversion fails, treat as 0 and add new value
                            new_value = value_num
                            self.log_activity(f"Warning: Existing value '{current_value}' was not numeric. Reset to {new_value}")
                    
                    # Update cell
                    cell.value = new_value
                    
                    # Log and record update
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row_idx}"
                    updated_cells.append(f"{data_type} at {cell_ref} (New value: {new_value})")
                    self.log_activity(f"Updated {data_type} for {trade}/{location}: {current_value} → {new_value}")
                    
                except Exception as e:
                    self.log_activity(f"Error updating {data_type}: {str(e)}")
                    continue
        
        # Save changes
        wb.save(file_path)
        wb.close()
        
        return updated_cells

    def process_input(self):
        """Handle button click event"""
        text = self.input_text.get("1.0", tk.END).strip()
        file_path = self.file_path.get().strip()
        
        if not text:
            self.log_activity("Error: Please enter input text")
            return
            
        if not file_path:
            self.log_activity("Error: Please select an Excel file")
            return
        
        if not os.path.exists(file_path):
            self.log_activity(f"Error: File not found: {file_path}")
            return
        
        try:
            data = self.parse_input(text)
            self.log_activity(f"Processing input: {', '.join(f'{k}={v}' for k, v in data.items())}")
            
            updated_cells = self.update_excel(
                file_path, 
                data,
                create_backup=self.backup_var.get()
            )
            
            if updated_cells:
                msg = f"Success! Updated {len(updated_cells)} cells"
                for cell in updated_cells:
                    msg += f"\n- {cell}"
                self.log_activity(msg)
                messagebox.showinfo("Success", msg)
            else:
                msg = "No cells updated. Check input parameters."
                self.log_activity(msg)
                messagebox.showinfo("Information", msg)
        except Exception as e:
            error_msg = f"Processing failed: {str(e)}"
            self.log_activity(error_msg)
            messagebox.showerror("Error", error_msg)

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelUpdaterApp(root)
    root.mainloop()